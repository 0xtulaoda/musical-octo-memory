from distutils.log import error
from seleniumbase import BaseCase
import os
from openpyxl import *
import random

the_address_PATH = os.path.abspath(
    r"..") + '/tools/Forge-export-tokenholders-for-contract-0x5198625a8abf34a0d2a1f262861ff3b3079302bf2.xlsx'

print('the_address_PATH', the_address_PATH)


class Excel:
    def __init__(self, file):
        self.file = file
        self.wb = load_workbook(self.file)
        # sheets = self.wb.get_sheet_names()
        # print(sheets, 'sheets')
        # sheets = self.wb['工作表 1']
        # self.sheet = sheets
        self.ws = self.wb['工作表 1 - Forge-export-tokenholde']

    def getColValues(self, column):
        rows = self.ws.max_row
        column_data = []
        for i in range(1, rows + 1):
            cell_value = self.ws.cell(row=i, column=column).value
            column_data.append(cell_value)
        return column_data

    def saveColValues(self, row, column, value):
        self.ws.cell(column, row, value)
        self.wb.save(the_address_PATH)


selector = '#PageWrapper > div > div:nth-child(1) > div.shared__HStack-sc-1qg837v-1.hAEXWs > div.shared__HStack-sc-1qg837v-1.iDsJUl > div.VStack-sc-1vdo21d-0.dDeGb > span > div > div.shared__HStack-sc-1qg837v-1.jZmJbE > div.UIText-sc-96tl0y-0.BreakWordText-sc-1s64evs-0.doIpXm.iHvGkp'
address_list = Excel(the_address_PATH).getColValues(1)[1:]
# balance_list = Excel(the_address_PATH).getColValues(2)

print('address_list', address_list)


class DemoException():  # DemoException(BaseException) 为正确写法
    """demo异常类型"""
    pass


# def demo_exc_handling():
#     print('-> coroutine started')
#     while True:
#         try:
#             var = yield
#         except DemoException as e:
#             print('*** DemoException handled. Continuing...')
#         else:
#             print('-> coroutine received: {!r}'.format(var))
#     # raise RuntimeError('This line should never run.')


# cor_exc = demo_exc_handling()
# cor_exc.send(None)
# cor_exc.send(1)


class TestMFALogin(BaseCase):
    def test_mfa_login(self):
        for index, address in enumerate(address_list):
            # self.wait(random.randint(0, 100))
            self.getData(address, index)
            # Excel(the_address_PATH).saveColValues(4, index+1, text)
    #     # self.save_screenshot_to_logs()
    # def test_mfa_login(self):
    #     for index, balance in enumerate(balance_list):
    #         # self.wait(random.randint(0, 100))
    #         # self.getData(address, index)
    #         print(balance, 'balance')
    #         try:
    #             Excel(the_address_PATH).saveColValues(
    #                 2, index+1, float(balance))
    #         except:
    #             print('*** DemoException handled. Continuing...')

    # self.save_screenshot_to_logs()

    def getData(self, address, index):
        print('index', index)
        url = 'https://app.zerion.io/'+address+'/nfts'
        self.open(url)
        try:
            self.wait_for_element(selector, timeout=30)
        except:
            print('*** DemoException handled. Continuing...')
        else:
            text = self.get_text(selector).replace('US$', '').replace(',', '')
            print(text)
            Excel(the_address_PATH).saveColValues(4, index+2, float(text))
